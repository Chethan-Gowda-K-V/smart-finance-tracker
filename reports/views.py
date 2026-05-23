import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from expenses.models import Transaction, Account, Category

# ReportLab Imports for PDF
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# OpenPyXL Imports for Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

@login_required
def export_pdf_view(request):
    user = request.user
    transactions = Transaction.objects.filter(user=user).select_related('category', 'account').order_by('-date')
    
    # Calculate summary metrics
    total_income = transactions.filter(type='income').aggregate(Sum('amount'))['amount__sum'] or 0.00
    total_expenses = transactions.filter(type='expense').aggregate(Sum('amount'))['amount__sum'] or 0.00
    net_savings = total_income - total_expenses
    
    # Create the PDF Buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#0f172a'), # Slate 900
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#64748b'), # Slate 500
        spaceAfter=20
    )
    section_title = ParagraphStyle(
        'SecTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#1e293b'), # Slate 800
        spaceBefore=15,
        spaceAfter=10
    )
    body_style = ParagraphStyle(
        'Body',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#334155')
    )
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )
    
    # Header Banner
    story.append(Paragraph("SMART FINANCE TRACKER", title_style))
    story.append(Paragraph(f"Financial Statement | Account: {user.username} | Generated on: {transactions.model.objects.none()._state.db or 'Live System'}", subtitle_style))
    story.append(Spacer(1, 10))
    
    # Summary Cards Table
    story.append(Paragraph("Executive Summary", section_title))
    summary_data = [
        [
            Paragraph("<b>Total Income</b>", body_style),
            Paragraph("<b>Total Expenses</b>", body_style),
            Paragraph("<b>Net Savings</b>", body_style)
        ],
        [
            f"{user.currency} {total_income:,.2f}",
            f"{user.currency} {total_expenses:,.2f}",
            f"{user.currency} {net_savings:,.2f}"
        ]
    ]
    summary_table = Table(summary_data, colWidths=[170, 170, 170])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 12),
        ('TEXTCOLOR', (0, 1), (0, 1), colors.HexColor('#10b981')), # Green for Income
        ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#ef4444')), # Red for Expense
        ('TEXTCOLOR', (2, 1), (2, 1), colors.HexColor('#3b82f6')), # Blue for Net
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Transactions List Table
    story.append(Paragraph("Transaction History", section_title))
    
    table_data = [[
        Paragraph("Date", header_style),
        Paragraph("Account", header_style),
        Paragraph("Category", header_style),
        Paragraph("Type", header_style),
        Paragraph("Description", header_style),
        Paragraph("Amount", header_style)
    ]]
    
    for tx in transactions:
        tx_type_formatted = Paragraph(f"<font color='{'green' if tx.type == 'income' else 'red'}'>{tx.type.upper()}</font>", body_style)
        table_data.append([
            Paragraph(tx.date.strftime('%Y-%m-%d'), body_style),
            Paragraph(tx.account.name, body_style),
            Paragraph(tx.category.name, body_style),
            tx_type_formatted,
            Paragraph(tx.description or "-", body_style),
            Paragraph(f"<b>{user.currency} {tx.amount:,.2f}</b>", body_style)
        ])
        
    tx_table = Table(table_data, colWidths=[70, 80, 85, 55, 130, 90])
    tx_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')), # Dark slate header
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
    ]))
    story.append(tx_table)
    
    # Build Document
    doc.build(story)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="smart_finance_statement_{user.username}.pdf"'
    response.write(pdf)
    return response

@login_required
def export_excel_view(request):
    user = request.user
    transactions = Transaction.objects.filter(user=user).select_related('category', 'account').order_by('-date')
    
    # Calculate summary metrics
    total_income = transactions.filter(type='income').aggregate(Sum('amount'))['amount__sum'] or 0.00
    total_expenses = transactions.filter(type='expense').aggregate(Sum('amount'))['amount__sum'] or 0.00
    net_savings = total_income - total_expenses
    
    wb = Workbook()
    
    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="1E293B")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    accent_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Write summary titles
    ws_summary["A1"] = "Smart Finance Tracker - Executive Summary"
    ws_summary["A1"].font = title_font
    ws_summary.merge_cells("A1:C1")
    ws_summary.row_dimensions[1].height = 25
    
    ws_summary["A3"] = "Metric"
    ws_summary["B3"] = "Amount"
    ws_summary["C3"] = "Currency"
    
    for col in ["A3", "B3", "C3"]:
        ws_summary[col].font = header_font
        ws_summary[col].fill = header_fill
        ws_summary[col].alignment = Alignment(horizontal="center")
        
    summary_rows = [
        ("Total Income", total_income, user.currency),
        ("Total Expenses", total_expenses, user.currency),
        ("Net Savings", net_savings, user.currency)
    ]
    
    for idx, (metric, amt, curr) in enumerate(summary_rows, start=4):
        ws_summary.cell(row=idx, column=1, value=metric).font = bold_font
        ws_summary.cell(row=idx, column=1).border = thin_border
        
        amt_cell = ws_summary.cell(row=idx, column=2, value=amt)
        amt_cell.font = bold_font
        amt_cell.number_format = "$#,##0.00" if user.currency == "USD" else "#,##0.00"
        amt_cell.border = thin_border
        
        if metric == "Total Income":
            amt_cell.font = Font(name="Calibri", size=11, bold=True, color="10B981")
        elif metric == "Total Expenses":
            amt_cell.font = Font(name="Calibri", size=11, bold=True, color="EF4444")
        else:
            amt_cell.font = Font(name="Calibri", size=11, bold=True, color="3B82F6")
            
        curr_cell = ws_summary.cell(row=idx, column=3, value=curr)
        curr_cell.font = normal_font
        curr_cell.alignment = Alignment(horizontal="center")
        curr_cell.border = thin_border
        
    # Column width auto-adjust
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # 2. Transactions Sheet
    ws_tx = wb.create_sheet(title="Transactions")
    ws_tx.views.sheetView[0].showGridLines = True
    
    headers = ["Date", "Account", "Category", "Type", "Description", "Amount"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_tx.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    ws_tx.row_dimensions[1].height = 20
    
    for row_idx, tx in enumerate(transactions, start=2):
        ws_tx.cell(row=row_idx, column=1, value=tx.date.strftime('%Y-%m-%d')).alignment = Alignment(horizontal="center")
        ws_tx.cell(row=row_idx, column=2, value=tx.account.name)
        ws_tx.cell(row=row_idx, column=3, value=tx.category.name)
        
        type_cell = ws_tx.cell(row=row_idx, column=4, value=tx.type.upper())
        type_cell.alignment = Alignment(horizontal="center")
        if tx.type == 'income':
            type_cell.font = Font(name="Calibri", size=11, bold=True, color="10B981")
        else:
            type_cell.font = Font(name="Calibri", size=11, bold=True, color="EF4444")
            
        ws_tx.cell(row=row_idx, column=5, value=tx.description or "")
        
        amt_cell = ws_tx.cell(row=row_idx, column=6, value=float(tx.amount))
        amt_cell.number_format = "$#,##0.00" if user.currency == "USD" else "#,##0.00"
        amt_cell.font = bold_font
        
        for c in range(1, 7):
            ws_tx.cell(row=row_idx, column=c).border = thin_border
            if row_idx % 2 == 1:
                ws_tx.cell(row=row_idx, column=c).fill = accent_fill
                
    for col in ws_tx.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_tx.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Write output to HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="smart_finance_sheet_{user.username}.xlsx"'
    wb.save(response)
    return response
